
# first program with MongoDB
from pymongo import MongoClient
import pprint
import bson
import openpyxl

def pipeline(year, unread):
    pipeline_2017_unread = [
        {"$project": {"_id": 0, "unread": 1, "date":1, "year": {"$year": "$date"},  "week": {"$week": "$date"} }},	
        {"$match": {"$and": [{"year": year}, {"unread": unread}]}},
        {"$group": {"_id": { "week":"$week"}, "count": {"$sum": 1}}},
    #    {"$group": {"_id": { "week":"$week", "unread": "$unread", "year":"$year"}, "count": {"$sum": 1}}},
        ]
    year_2017_unread = list(emails.aggregate(pipeline_2017_unread))
    
    #pprint.pprint(year_2017_unread)
    
    pprint.pprint(year_2017_unread[1]["_id"])
    
    year_2017_unread_weeks = {}
    for week in year_2017_unread:
        pprint.pprint(week)
        year_2017_unread_weeks[week["_id"]["week"]] = week["count"]
    
    return year_2017_unread_weeks

client = MongoClient()
db = client.myEmails
emails = db.myEmailsTest6
pprint.pprint(emails.find_one())
#for email in emails.find():
#    pprint.pprint(email)

year_2017_unread_weeks = pipeline(2017, True)
weeks_unread = list(year_2017_unread_weeks.keys())
weeks_unread.sort()
pprint.pprint(year_2017_unread_weeks)

year_2017_read_weeks = pipeline(2017, False)
weeks_read = list(year_2017_read_weeks.keys())
weeks_read.sort()
pprint.pprint(year_2017_read_weeks)

wb = openpyxl.load_workbook("C:\\Users\\gseive\\Documents\\My emails\\project my_emails.xlsx")
wb.get_sheet_names()
sheet = wb.active

for i, w in enumerate(weeks_read):
    sheet.cell(row=2,column=i+2).value = year_2017_read_weeks[w]
for i, w in enumerate(weeks_unread):
    sheet.cell(row=3,column=i+2).value = year_2017_unread_weeks[w]

wb.save("C:\\Users\\gseive\\Documents\\My emails\\project my_emails_output.xlsx")
